import requests
import openpyxl
import urllib
import xml.etree.ElementTree as ET
import re
from fuzzywuzzy import fuzz
from operator import itemgetter
from datetime import datetime

# Functie om trefwoorden te reconciliëren met de Getty API op basis van AAT-zoekopdrachten
def reconcile_AAT_keyword(keyword):
    url = "http://vocabsservices.getty.edu/AATService.asmx/AATGetTermMatch?term="
    try:
        response = requests.get(url + urllib.parse.quote(keyword) + '&logop=and&notes=')
        results = ET.fromstring(response.content)
        resources = []
        for child in results.iter('Preferred_Parent'):
            try:
                name = re.sub(r'\[.+?\]', '', child.text.split(',')[0]).strip()
                page_link = re.search(r"\[(.+?)\]", child.text.split(',')[0]).group(1)
                uri = "http://vocab.getty.edu/page/aat/" + page_link
                score = fuzz.token_sort_ratio(keyword, name)
                match = score > 95
                resource = {
                    "uri": uri,
                    "name": name,
                    "score": score,
                    "match": match
                }
                resources.append(resource)
            except AttributeError:
                pass
        sorted_resources = sorted(resources, key=itemgetter('score'), reverse=True)
        return sorted_resources[:10]
    except Exception as e:
        print("Error:", e)
        return None

# Functie om het Excel-bestand te lezen en een nieuw bestand te maken met gereconcilieerde URI's en trefwoorden
def reconcile_and_export_excel(input_file_path, output_file_path, column_name):
    workbook = openpyxl.load_workbook(input_file_path)
    sheet = workbook.active
    max_row = sheet.max_row

    column_index = None
    for cell in sheet[1]:
        if cell.value == column_name:
            column_index = cell.column
            break

    if column_index is None:
        print(f"Kolom '{column_name}' niet gevonden in het Excel-bestand.")
        return

    new_workbook = openpyxl.Workbook()
    new_sheet = new_workbook.active

    new_sheet.cell(row=1, column=1, value=column_name)
    new_sheet.cell(row=1, column=2, value="Gereconcilieerde URI")
    new_sheet.cell(row=1, column=3, value="Gereconcilieerde Naam")

    for row_num in range(2, max_row + 1):
        keyword = sheet.cell(row=row_num, column=column_index).value
        new_sheet.cell(row=row_num, column=1, value=keyword)
        if keyword:
            reconciled_resources = reconcile_AAT_keyword(keyword)
            if reconciled_resources:
                new_sheet.cell(row=row_num, column=2, value=reconciled_resources[0]["uri"])
                new_sheet.cell(row=row_num, column=3, value=reconciled_resources[0]["name"])
            else:
                new_sheet.cell(row=row_num, column=2, value="Geen match gevonden")
                new_sheet.cell(row=row_num, column=3, value="Geen match gevonden")
        else:
            new_sheet.cell(row=row_num, column=2, value="Geen trefwoord")
            new_sheet.cell(row=row_num, column=3, value="Geen trefwoord")

        # Voortgangsindicator
        print(f"Verwerken van rij {row_num - 1} van {max_row - 1} ({(row_num - 1) / (max_row - 1) * 100:.2f}%)")

    # Datum en tijd toevoegen aan de uitvoerbestandsnaam
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    sanitized_column_name = re.sub(r'\W+', '_', column_name)  # Replaces non-alphanumeric characters with underscores
    output_file_path_with_timestamp = output_file_path.replace(".xlsx", f"_{sanitized_column_name}_{timestamp}.xlsx")

    new_workbook.save(output_file_path_with_timestamp)
    print(f"Bestand opgeslagen als: {output_file_path_with_timestamp}")

# Pad naar het invoer- en uitvoerbestand
input_excel_file_path = "/Users/patrickmout/Downloads/VGWW/Onderzoek-Van-Gogh-tbv-Van-Gogh-Worldwide_AAT.xlsx"
output_excel_file_path = "/Users/patrickmout/Downloads/VGWW/analysemethode_match_AAT.xlsx"
# Kolomnaam waarin trefwoorden worden opgeslagen
column_name = "trefwoord.inhoud 8"

# Reconcilieer het Excel-bestand en exporteer naar een nieuw bestand
reconcile_and_export_excel(input_excel_file_path, output_excel_file_path, column_name)
